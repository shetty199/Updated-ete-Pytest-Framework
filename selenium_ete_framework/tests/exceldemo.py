# import openpyxl
#
# book=openpyxl.load_workbook("C:\\Users\\shettappa.chalawadi\\Downloads\\Demo1.xlsx")
# sheet=book.active
# cell=sheet.cell(row=2,column=1)
# print(cell.value)


import openpyxl
book=openpyxl.load_workbook("C:\\Users\\shettappa.chalawadi\\Downloads\\Demo1.xlsx")
sheet=book.active
# cell=sheet.cell(row=2,column=1)
# print(cell.value)

# sheet.cell(row=4,column=1).value='chalawadi'
#
# print(sheet.cell(row=4,column=1).value)
#
# print(sheet['A3'].value)

for i in range(1,sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i, column=j).value)
    print(sheet.cell(row=i,column=1).value)